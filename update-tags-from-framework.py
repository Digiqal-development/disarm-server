import json
import openpyxl

PHASES = ['Plan', 'Prepare', 'Execute', 'Asses']

TACTICS = {
    "TA01": { "title": "Plan Strategy", "phase": "Plan", "rank": 1 },
    "TA02": { "title": "Plan Objectives", "phase": "Plan", "rank": 2 },
    "TA13": { "title": "Target Audience Analysis", "phase": "Plan", "rank": 3 },
    "TA14": { "title": "Develop Narratives", "phase": "Prepare", "rank": 4 },
    "TA06": { "title": "Develop Content", "phase": "Prepare", "rank": 5 },
    "TA15": { "title": "Establish Assets", "phase": "Prepare", "rank": 6 },
    "TA16": { "title": "Establish Legitimacy",  "phase": "Prepare", "rank": 7 },
    "TA05": { "title": "Microtarget", "phase": "Prepare", "rank": 8 },
    "TA07": { "title": "Select Channels and Affordances", "phase": "Prepare", "rank": 9 },
    "TA08": { "title": "Conduct Pump Priming", "phase": "Execute", "rank": 10 },
    "TA09": { "title": "Deliver Content", "phase": "Execute",  "rank": 11 },
    "TA17": { "title": "Maximise Exposure", "phase": "Execute", "rank": 12 },
    "TA18": { "title": "Drive Online Harms", "phase": "Execute", "rank": 13 },
    "TA10": { "title": "Drive Offline Activity", "phase": "Execute", "rank": 14 },
    "TA11": { "title": "Persist in the Information Environment", "phase": "Execute", "rank": 15 },
    "TA12": { "title": "Assess Effectiveness", "phase": "Assess", "rank": 16  } }


def get_use_name(id):
     return TACTICS[id]["title"]

def get_use_phase(id):
     return TACTICS[id]["phase"]




workbook = openpyxl.load_workbook('DISARM_FRAMEWORKS_MASTER.xlsx')
sheet = workbook.get_sheet_by_name('techniques')
max_row = sheet.max_row
max_column = sheet.max_column

tags_data = {"ids": {}}

try:
    for row in range(2, max_row + 1):
            id = sheet.cell(row=row, column=1).value
            title = sheet.cell(row=row, column=2).value

            use_id = sheet.cell(row=row, column=4).value
            use = get_use_name(use_id)
            phase = get_use_phase(use_id)

            description = sheet.cell(row=row, column=5).value

            tags_data["ids"][id] = {
                "title": title,
                "use": use + " [" + use_id + "]",
                "phase": phase,
                "description": description
            }
except Exception as e: 
    print(e)


with open('tags.json', 'w') as tags_file:
    json.dump(tags_data, tags_file, indent=2)







    